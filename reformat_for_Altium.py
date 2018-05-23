##########################################
## Re-formatting netlist for Altium import
##########################################

import openpyxl
import csv
import sys
from operator import itemgetter

### Select print option ###
bool_print_DCB = True
bool_print_PT = False
bool_print_Pathfinder = False
###########################


file_SEAM = "D:/UT/Backplane-Remapping/trueType/backplaneMapping_SEAMPins_trueType_v4.1.xlsm"
# Load the xlsm workbook
book_SEAM = openpyxl.load_workbook(file_SEAM)

## Reading SEAM file: from Column B to K, from Row 6 to 405
### Header format:
#print('SEAM slot'+','+'SEAM pin'+','+'Signal ID'+','+'Pigtail slot'+','+'Pigtail pin'+','+'GBTx ID'+','+'Conn to PT'+','+'DCB slot'+','+'DCB pin'+','+'Conn to other DCBs')

listDCB = []

for num_tab in range(0,12):
    ## read DCB slots  
    tab = book_SEAM.get_sheet_by_name(str(num_tab))
    listTab = []
    for num_row in range(6,406):
        listEntry = []
        
        ## read 400 pins from row 6 to 405
        ### format strings for referencing cells
        str_ref = 'B'+str(num_row)
        str_SmPin = 'C'+str(num_row)
        str_SigID = 'D'+str(num_row)
        str_PtSlot = 'E'+str(num_row)
        str_PtPin = 'G'+str(num_row)
        str_GBTxID = 'H'+str(num_row)
        str_DcbSlot = 'J'+str(num_row)
        str_DcbPin = 'K'+str(num_row)

        ## Filling the lists: DCB##, X, ##, PT/DCB##, X, ##, ID
        if ( str(tab[str_SigID].value) != 'AGND' ):
            if ( str(tab[str_PtPin].value) != 'None' ): #DCB-PT
                listEntry.append(str(num_tab))
                listEntry.append( str(tab[str_SmPin].value)[0])
                if int(str(tab[str_SmPin].value[1:]))< 10: #for single-digi number use 0X
                    listEntry.append( '0'+str(tab[str_SmPin].value)[1:])
                else:
                    listEntry.append( str(tab[str_SmPin].value)[1:])
                listEntry.append(str(int(str.split(str(tab[str_PtSlot].value))[0])) )
                listEntry.append(str(tab[str_PtPin].value)[0])
                listEntry.append(str(tab[str_PtPin].value)[1:])
                listEntry.append(str(tab[str_SigID].value))
            if ( str(tab[str_PtPin].value) == 'None' ):
                if (str(tab[str_DcbPin].value) != 'None'): #DCB-DCB
                    listEntry.append(str(num_tab))
                    listEntry.append( str(tab[str_SmPin].value)[0])
                    if int(str(tab[str_SmPin].value[1:]))< 10:
                        listEntry.append( '0'+str(tab[str_SmPin].value)[1:])
                    else:
                        listEntry.append( str(tab[str_SmPin].value)[1:])
                    listEntry.append(str(int( str.split( str(tab[str_DcbSlot].value))[0])) )
                    listEntry.append( str(tab[str_DcbPin].value)[0])
                    if int(str(tab[str_DcbPin].value[1:]))< 10:
                        listEntry.append( '0'+str(tab[str_DcbPin].value)[1:])
                    else:
                        listEntry.append( str(tab[str_DcbPin].value)[1:])
                    listEntry.append(str(tab[str_SigID].value))
                    ### 7th entry for DCB-DCB###
                    listEntry.append("DCB-DCB") 
                else:   #DCB - nothing
                    listEntry.append(str(num_tab))
                    listEntry.append( str(tab[str_SmPin].value)[0])
                    if int(str(tab[str_SmPin].value[1:]))< 10:
                        listEntry.append( '0'+str(tab[str_SmPin].value)[1:])
                    else:
                        listEntry.append( str(tab[str_SmPin].value)[1:])
                    listEntry.append('None')
                    listEntry.append('None')
                    listEntry.append('None')
                    listEntry.append(str(tab[str_SigID].value))
        if ( str(tab[str_SigID].value) == 'AGND' ):
            if ( str(tab[str_PtPin].value) != 'None' ): #DCB-PT
                listEntry.append(str(num_tab))
                listEntry.append( str(tab[str_SmPin].value)[0])
                if int(str(tab[str_SmPin].value[1:]))< 10:
                    listEntry.append( '0'+str(tab[str_SmPin].value)[1:])
                else:
                    listEntry.append( str(tab[str_SmPin].value)[1:])
                listEntry.append(str(tab[str_PtSlot].value))
                listEntry.append('')
                listEntry.append(str(tab[str_PtPin].value)) ##mult. pins--careful here
                listEntry.append(str(tab[str_SigID].value))
            if ( str(tab[str_PtPin].value) == 'None' ): #DCB-nothing
               
                listEntry.append(str(num_tab))
                listEntry.append( str(tab[str_SmPin].value)[0])
                if int(str(tab[str_SmPin].value[1:]))< 10:
                    listEntry.append( '0'+str(tab[str_SmPin].value)[1:])
                else:
                    listEntry.append( str(tab[str_SmPin].value)[1:])
                listEntry.append('None')
                listEntry.append('None')
                listEntry.append('None')
                listEntry.append(str(tab[str_SigID].value))

        listTab.append(listEntry)
## Now sort listTotal according to DCB#X##
#    listTab.sort()
    listTab.sort(key=itemgetter(1,2))
    listDCB.append(listTab)

##for aa in range(0, len(listDCB[0])):
##    net = listDCB[0][aa]
##    print(net)
    #print(net[0]+net[1]+net[2]+'.'+net[3]+net[4]+str(net[5])+'.'+net[6])



########################################################
## Now from PT side:

file_PT = "D:/UT/Backplane-Remapping/trueType/backplaneMapping_pigtailPins_trueType_strictDepopulation_v5.1.xlsm"
book_PT = openpyxl.load_workbook(file_PT)
listPT = []

for num_tab in range(0,12):
    ## read PT slots
    tab = book_PT.get_sheet_by_name(str(num_tab))
    listTab = []
    for num_row in range(6,406):
        listEntry = []
        ## read 400 pins from row 6 to 405
        ### format strings for referencing cells
        str_ref = 'B'+str(num_row)
        str_PtPin = 'C'+str(num_row)
        str_SigID = 'D'+str(num_row)
        str_DcbSlot = 'E'+str(num_row)
        str_DcbPin = 'G'+str(num_row)
        str_GBTxID = 'H'+str(num_row)
        ## Filling the lists: DCB##, X, ##, PT##, X, ##, ID
        if ( str(tab[str_DcbPin].value) != 'None' ): #DCB-PT
            listEntry.append( str(int(str.split(str(tab[str_DcbSlot].value))[0])) )
            listEntry.append( str(tab[str_DcbPin].value)[0])
            if int(str(tab[str_DcbPin].value[1:]))< 10: #for single-digi number use 0X
                listEntry.append( '0'+str(tab[str_DcbPin].value)[1:])
            else:
                listEntry.append( str(tab[str_DcbPin].value)[1:])
            listEntry.append(str(num_tab))
            listEntry.append(str(tab[str_PtPin].value)[0])
            listEntry.append(str(tab[str_PtPin].value)[1:])
            listEntry.append(str(tab[str_SigID].value))
        if ( str(tab[str_DcbPin].value) == 'None' ): #PT - nothing            
            listEntry.append('None')
            listEntry.append( 'None' )
            listEntry.append( 'None' )
            listEntry.append(str(num_tab))
            listEntry.append(str(tab[str_PtPin].value)[0])
            listEntry.append(str(tab[str_PtPin].value)[1:])
            listEntry.append(str(tab[str_SigID].value))

        listTab.append(listEntry)
#### Now sort listPT according to PT#X##
    listTab.sort(key=itemgetter(4,5))
    listPT.append(listTab)

##for aa in range(0, len(listPT[0])):
##    net = listPT[0][aa]
##    print(net)

#### Now replace the Signal ID by finding the DCB side net:
for ptX in range(0,len(listPT)):
    ptSublist = listPT[ptX]
    for i in range(0,len(ptSublist)):
        ptNet = ptSublist[i]
        if ptNet[0]=='None': #if PT-nothing
            continue
        dcbX = int(ptNet[0]) #####!!!!! To be updated when generating full netlist (this works for pathfinder)
        for ii in range(0,len(listDCB[dcbX])):
            dcbNet = listDCB[dcbX][ii]
            if (ptNet[1] == dcbNet[1]) and (ptNet[3]==dcbNet[3]) and ( (ptNet[4]+ptNet[5])==(dcbNet[4]+dcbNet[5])):
               listPT[ptX][i][6] = dcbNet[6]


#### Now print out the reformated netlist
# but first get the list for telemitry breakout boards' net names
file_BOB = "D:/UT/Backplane-Remapping/Pathfinder/BrkOutBrd_Pin_Assignments_Mar27_2018_PM1.xlsx"
book_PT = openpyxl.load_workbook(file_BOB)
pinout_BOB = book_PT.get_sheet_by_name('PinAssignments')
list_BOB = []
for num_row in range(4,19):
    str_rowA = 'A'+str(num_row)
    str_rowD = 'D'+str(num_row)
    str_rowF = 'F'+str(num_row)
    str_rowI = 'I'+str(num_row)
    list_BOB.append( str(pinout_BOB[str_rowA].value) )
    list_BOB.append( str(pinout_BOB[str_rowD].value) )    
    list_BOB.append( str(pinout_BOB[str_rowF].value) )    
    list_BOB.append( str(pinout_BOB[str_rowI].value) )    
for num_row in range(55,70):
    str_rowA = 'A'+str(num_row)
    str_rowD = 'D'+str(num_row)
    str_rowF = 'F'+str(num_row)
    str_rowI = 'I'+str(num_row)
    list_BOB.append( str(pinout_BOB[str_rowA].value) )
    list_BOB.append( str(pinout_BOB[str_rowD].value) )    
    list_BOB.append( str(pinout_BOB[str_rowF].value) )    
    list_BOB.append( str(pinout_BOB[str_rowI].value) )    
for num_row in range(106,121):
    str_rowA = 'A'+str(num_row)
    str_rowD = 'D'+str(num_row)
    str_rowF = 'F'+str(num_row)
    str_rowI = 'I'+str(num_row)
    list_BOB.append( str(pinout_BOB[str_rowA].value) )
    list_BOB.append( str(pinout_BOB[str_rowD].value) )    
    list_BOB.append( str(pinout_BOB[str_rowF].value) )    
    list_BOB.append( str(pinout_BOB[str_rowI].value) )
for num_row in range(4,54):
    str_rowK = 'K'+str(num_row)
    str_rowN = 'N'+str(num_row)
    list_BOB.append( str(pinout_BOB[str_rowK].value) )
    list_BOB.append( str(pinout_BOB[str_rowN].value) )
for num_row in range(55,105):
    str_rowK = 'K'+str(num_row)
    str_rowN = 'N'+str(num_row)
    list_BOB.append( str(pinout_BOB[str_rowK].value) )
    list_BOB.append( str(pinout_BOB[str_rowN].value) )
for num_row in range(106,156):
    str_rowK = 'K'+str(num_row)
    str_rowN = 'N'+str(num_row)
    list_BOB.append( str(pinout_BOB[str_rowK].value) )
    list_BOB.append( str(pinout_BOB[str_rowN].value) )  
#print('len of list_BOB ', len(list_BOB))


#   0     1   2    3       4   5   6
# {DCB##, X, ##, PT/DCB##, X, ##, ID}
if bool_print_DCB:
    for i in range(0,len(listDCB)):
        if bool_print_Pathfinder and (i not in [0,2,4]): continue # skip tabs for pathfinder
        for ii in range(0, len(listDCB[i])):
            net = listDCB[i][ii]
            if (net[6]=='AGND'): #for AGND
                print('JD'+net[0]+'_AGND'+ ',' +net[0]+','+net[1]+net[2]+','+net[3]+','+net[4]+net[5])
                continue
            if (net[5]=='None'):
                if net[6] not in ['GND','1.5V','2.5V','1V5_SENSE_NEG','1V5_SENSE_POS']:
                    print('JD'+net[0]+net[1]+net[2]+'_ForRefOnly_'+net[6]+ ',' +net[0]+','+net[1]+net[2]+','+''+','+'')
                elif net[6] == 'GND':
                    print('JD'+net[0]+'_GND'+ ',' +net[0]+','+net[1]+net[2]+','+''+','+'')
                elif net[6]=='1.5V':
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue
                        if ( ('JD'+net[0]) in list_BOB[a].split('_') ) and ( '1V5' in list_BOB[a] ) and ( 'SENSE' not in list_BOB[a] ):
                            print(list_BOB[a][:-2] + ',' +net[0]+','+net[1]+net[2]+','+''+','+'')
                            break # so not to be duplicated by 1V5_M/S 
                elif net[6]=='2.5V':
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue ##For 2V5, 2 JDs share one net name
                        if ( '2V5' in list_BOB[a] ) and ( ('JD'+net[0])==list_BOB[a].split('_')[0] or net[0]==list_BOB[a].split('_')[1]) and ( 'SENSE' not in list_BOB[a] ):
                            print(list_BOB[a] + ',' +net[0]+','+net[1]+net[2]+','+''+','+'')
                elif net[6]=='1V5_SENSE_NEG':
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue
                        if ( ('JD'+net[0]) in list_BOB[a].split('_') ) and ( '1V5_SENSE_N' in list_BOB[a] ) :
                            print(list_BOB[a] + ',' +net[0]+','+net[1]+net[2]+','+''+','+'')
                elif net[6]=='1V5_SENSE_POS':
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue
                        if ( ('JD'+net[0]) in list_BOB[a].split('_') ) and ( '1V5_SENSE_P' in list_BOB[a] ) :
                            print(list_BOB[a] + ',' +net[0]+','+net[1]+net[2]+','+''+','+'')
                    
            elif (len(net)== 7): #len=7 for DCB-PT
                print('JD'+net[0]+net[1]+net[2]+'_JP'+net[3]+net[4]+net[5]+'_'+net[6]+ ',' +net[0]+','+net[1]+net[2]+','+net[3]+','+net[4]+net[5])
            elif (len(net)== 8): #len=8 for DCB-DCB
                if int(net[0])<int(net[3]):
                    print('JD'+net[0]+net[1]+net[2]+'_JD'+net[3]+net[4]+net[5]+'_'+net[6]+ ',' +net[0]+','+net[1]+net[2]+','+net[3]+','+net[4]+net[5])
                if int(net[0])>int(net[3]):
                    print('JD'+net[3]+net[4]+net[5]+'_JD'+net[0]+net[1]+net[2]+'_'+net[6]+ ',' +net[0]+','+net[1]+net[2]+','+net[3]+','+net[4]+net[5])

#   0     1   2    3   4   5   6
# {DCB##, X, ##, PT##, X, ##, ID}
if bool_print_PT:
    for i in range(0,len(listPT)):
        for ii in range(0, len(listPT[i])):
            net = listPT[i][ii]
            if bool_print_Pathfinder and (i not in [0,1]):# for pathfinder, skip non-BOB nets when not in slot 0/1
                if ('LV_SOURCE' not in net[6]) and ('LV_RETURN' not in net[6]) and ('LV_SENSE' not in net[6]) and ('THERMISTOR' not in net[6]):
                    print(''+ ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                    continue
            if (net[2]=='None'):
                if ('LV_SOURCE' not in net[6]) and ('LV_RETURN' not in net[6]) and ('LV_SENSE' not in net[6]) and ('THERMISTOR' not in net[6]):
                    print('JP'+net[3]+net[4]+net[5]+'_ForRefOnly_'+net[6]+ ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                elif ('LV_SOURCE' in net[6]):
                    replaced = False
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue
                        if ( ('JP'+net[3]) in list_BOB[a].split('_') ) and ( net[6] in list_BOB[a] ) :
                            print(list_BOB[a] + ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                            replaced=True
                    if replaced==False:
                        print('JP'+net[3]+net[4]+net[5]+'_ForRefOnly_'+net[6]+ ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                elif ('LV_RETURN' in net[6]):
                    replaced = False
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue
                        if ( ('JP'+net[3]) in list_BOB[a].split('_') ) and ( net[6] in list_BOB[a] ):
                            print(list_BOB[a] + ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                            replaced=True
                    if replaced==False:
                        print('JP'+net[3]+net[4]+net[5]+'_ForRefOnly_'+net[6]+ ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                elif ('LV_SENSE' in net[6]):
                    replaced = False
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue
                        if ( ('JP'+net[3]) in list_BOB[a].split('_') ) and ( net[6] in list_BOB[a] ):
                            print(list_BOB[a] + ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                            replaced=True
                    if replaced==False:
                        print('JP'+net[3]+net[4]+net[5]+'_ForRefOnly_'+net[6]+ ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                elif ('THERMISTOR' in net[6]):
                    replaced = False
                    for a in range(0,len(list_BOB)):
                        if list_BOB[a]=='GND' or list_BOB[a]=='': continue
                        if ( ('JP'+net[3]) in list_BOB[a].split('_') ) and ( net[6] in list_BOB[a] ):
                            print(list_BOB[a] + ',' +net[3]+','+net[4]+net[5]+','+''+','+'')
                            replaced = True
                    if replaced == False: # if not replaced by the name 
                        print('JP'+net[3]+net[4]+net[5]+'_ForRefOnly_'+net[6]+ ',' +net[3]+','+net[4]+net[5]+','+''+','+'')                            
            else: # PT-DCB
                print('JD'+net[0]+net[1]+net[2]+'_JP'+net[3]+net[4]+net[5]+'_'+net[6]+ ',' +net[3]+','+net[4]+net[5]+','+''+','+'')


##            input("Press 0 to continue...")
                

